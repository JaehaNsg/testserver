import socket
from servermodule.python_path_config import *
import openpyxl

def read_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        key = row[0]
        data.append(row)
    return data

file_path = "/Users/jeongjaehan/Desktop/testdata.xlsx"
excel_data = read_excel_data(file_path)
formatted_data = '\n'.join([', '.join(map(str, row)) for row in excel_data])

# 서버 설정: IP 주소와 포트 번호
def get_ip_address():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # 작동하지 않는 IP 주소인 경우도 있지만 이 경로를 사용하여
        # 로컬 IP 주소를 찾는 것에는 문제가 없습니다.
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

server_ip = get_ip_address()
print(server_ip)
server_port = 1103

# 소켓 객체 생성 및 설정
server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
server_socket.bind((server_ip, server_port))

# 클라이언트 연결 대기
server_socket.listen(1)
print(f"서버가 {server_ip}:{server_port}에서 클라이언트 연결을 대기합니다...")

# 클라이언트 연결 수락
client_socket, client_addr = server_socket.accept()
print(f"클라이언트 {client_addr}가 연결되었습니다.")

# 클라이언트로부터 데이터 수신
data = client_socket.recv(1024).decode('utf-8')
print(f"클라이언트로부터 받은 데이터: {data}")

# 클라이언트에게 데이터 전송
client_socket.send("안녕하세요, 클라이언트!".encode('utf-8'))
client_socket.send(formatted_data.encode('utf-8'))

# 클라이언트로부터 키워드 수신
while True:
    keyword = client_socket.recv(1024).decode('utf-8')
    if keyword == "quit":
        break
    match = excel_data.get(keyword, None)
    if match is None:
        client_socket.send("No match found\n".encode('utf-8'))
    else:
        formatted_data = ', '.join(map(str, match))
        client_socket.send(f"{formatted_data}\n".encode('utf-8'))

# 소켓 종료
client_socket.close()
server_socket.close()
