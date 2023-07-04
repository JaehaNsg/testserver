import sys

def path_finder(): #pip 경로 설정
    sys.path.append('/opt/anaconda3/bin')
    sys.path.append('/opt/anaconda3/lib/python3.9/site-packages')
    
path_finder()

import openpyxl

def read_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        key = row[0]
        data.append(row)
    return data

file_path = "/Users/jeongjaehan/Desktop/testdata.xlsx" #파일 경로 설정
excel_data = read_excel_data(file_path)
formatted_data = '\n'.join([', '.join(map(str, row)) for row in excel_data])